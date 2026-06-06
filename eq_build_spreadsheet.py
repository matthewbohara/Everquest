"""
EverQuest Tradeskill 350 Spreadsheet Builder
=============================================
Builds a personalized Excel spreadsheet from your scraped recipe data
and inventory file. No AI or external tools required.

REQUIRES:
  - scraped_ingredients/ingredients_WITH_SOURCES.csv  (from eq_ingredient_sources.py)
  - Your EQ inventory outputfile: CharName_Server-Inventory.txt

SETUP:  pip install openpyxl
USAGE:  python eq_build_spreadsheet.py

OUTPUT: EQ_Tradeskill_350_Tracker.xlsx
"""

import csv, os, sys, re, glob
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Run: pip install openpyxl")
    sys.exit(1)

SCRIPT_FOLDER   = os.path.dirname(os.path.abspath(__file__))
SOURCES_CSV     = os.path.join(SCRIPT_FOLDER, 'scraped_ingredients', 'ingredients_WITH_SOURCES.csv')
OUTPUT_FILE     = os.path.join(SCRIPT_FOLDER, 'EQ_Tradeskill_350_Tracker.xlsx')

NEEDED_FOR_350 = {
    'Alchemy': 562, 'Baking': 698, 'Blacksmithing': 1939, 'Brewing': 302,
    'Fishing': 122, 'Fletching': 876, 'Jewelry Making': 1279,
    'Make Poison': 543, 'Pottery': 2014, 'Research': 2570,
    'Tailoring': 2078, 'Tinkering': 791,
}

SKILL_ORDER = [
    'Brewing', 'Fishing', 'Alchemy', 'Baking', 'Make Poison',
    'Tinkering', 'Fletching', 'Jewelry Making', 'Blacksmithing',
    'Tailoring', 'Pottery', 'Research'
]

# ── STYLES ────────────────────────────────────────────────────────────────────
thin  = Side(style='thin', color='BFBFBF')
BDR   = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR   = PatternFill('solid', start_color='1F4E79')
SUB   = PatternFill('solid', start_color='2E75B6')
GRN   = PatternFill('solid', start_color='E2EFDA')
GRN2  = PatternFill('solid', start_color='D5E8D4')
YLW   = PatternFill('solid', start_color='FFF2CC')
YLW2  = PatternFill('solid', start_color='FFF9E6')
RED   = PatternFill('solid', start_color='FCE4D6')
RED2  = PatternFill('solid', start_color='F9D7CD')
ALT   = PatternFill('solid', start_color='F2F2F2')
WHT   = PatternFill('solid', start_color='FFFFFF')
DGRN  = PatternFill('solid', start_color='375623')
DORG  = PatternFill('solid', start_color='843C0C')
DPURP = PatternFill('solid', start_color='4B0082')

def sc(cell, bold=False, color='000000', size=10, fill=WHT,
       align='left', italic=False, wrap=True):
    cell.font = Font(name='Arial', bold=bold, color=color, size=size, italic=italic)
    cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border = BDR

# ── LOAD INVENTORY ────────────────────────────────────────────────────────────
def find_inventory_file():
    """Auto-detect EQ inventory outputfile."""
    patterns = ['*-Inventory.txt', '*-inventory.txt']
    for pattern in patterns:
        matches = glob.glob(os.path.join(SCRIPT_FOLDER, pattern))
        if matches:
            # Return most recently modified
            return max(matches, key=os.path.getmtime)
    return None

def load_inventory(filepath):
    """Load inventory into dict of {item_name_lower: quantity}."""
    inventory = defaultdict(int)
    if not filepath or not os.path.exists(filepath):
        return inventory
    with open(filepath, encoding='utf-8', errors='replace') as f:
        lines = f.readlines()
    for line in lines[1:]:  # skip header
        parts = line.strip().split('\t')
        if len(parts) >= 4:
            name = parts[1].strip()
            try:
                qty = int(parts[3])
            except:
                qty = 1
            if name and name != 'Empty':
                inventory[name.lower()] += qty
    print(f"  Loaded {len(inventory)} unique items from inventory")
    return inventory

# ── LOAD RECIPES ──────────────────────────────────────────────────────────────
def load_recipes():
    """Load ingredients_WITH_SOURCES.csv."""
    if not os.path.exists(SOURCES_CSV):
        print(f"ERROR: {SOURCES_CSV} not found!")
        print("Run eq_ingredient_sources.py first.")
        sys.exit(1)

    recipes = defaultdict(list)
    with open(SOURCES_CSV, encoding='utf-8') as f:
        for row in csv.DictReader(f):
            skill = row.get('skill', '').strip()
            if skill:
                recipes[skill].append(row)

    total = sum(len(v) for v in recipes.values())
    print(f"  Loaded {total} recipes across {len(recipes)} skills")
    return recipes

# ── INGREDIENT STATUS ─────────────────────────────────────────────────────────
def ingredient_status(ingr_name, ingr_src, inventory):
    """Returns (have_it, status_text, fill_color)"""
    if not ingr_name:
        return None, '', WHT

    clean = re.sub(r'\s*\(\d+\)\s*$', '', ingr_name).strip().lower()
    qty = inventory.get(clean, 0)

    if qty > 0:
        return True, f'✓ Have ({qty})', GRN
    elif 'vendor' in ingr_src.lower():
        return False, 'Buy from vendor', YLW
    elif 'foraged' in ingr_src.lower():
        return False, 'Forage', RED
    elif 'fished' in ingr_src.lower():
        return False, 'Fish', RED
    elif 'dropped' in ingr_src.lower():
        return False, 'Farm/Drop', RED2
    elif 'crafted' in ingr_src.lower():
        return False, 'Subcombine', YLW2
    else:
        return False, 'Check EQTraders', ALT

def recipe_overall_status(row, inventory):
    """Determine overall recipe status based on ingredients."""
    all_have = True
    any_farmable = False

    for i in range(1, 9):
        ingr = row.get(f'ingredient_{i}', '').strip()
        if not ingr:
            continue
        src  = row.get(f'src_{i}', '').strip()
        have, _, _ = ingredient_status(ingr, src, inventory)
        if not have:
            all_have = False
            if 'vendor' in src.lower() or 'subcombine' in src.lower():
                pass
            else:
                any_farmable = True

    if all_have:
        return 'Ready Now!', GRN
    elif not any_farmable:
        return 'Buy/Craft Only', YLW
    else:
        return 'Need to Farm', RED

# ── BUILD SPREADSHEET ─────────────────────────────────────────────────────────
def build_spreadsheet(recipes, inventory):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── SUMMARY SHEET ─────────────────────────────────────────────────────────
    ws = wb.create_sheet('Summary')
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14

    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value = 'EverQuest Tradeskill 350 — Missing Recipe Tracker'
    sc(c, bold=True, color='FFFFFF', size=13, fill=HDR, align='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:H2')
    c = ws['A2']
    c.value = ('Green = Ready to make now  |  Yellow = Buy/vendor ingredients only  |  '
               'Red = Need to farm ingredients  |  Click skill name to jump to that tab')
    sc(c, italic=True, size=9, color='595959', fill=YLW)
    ws.row_dimensions[2].height = 18

    hdrs = ['Tradeskill', 'Missing Recipes', 'Need for 350', '% to Goal',
            'Ready Now!', 'Buy Only', 'Need Farm', 'Still Need']
    for col, hdr in enumerate(hdrs, 1):
        c = ws.cell(row=3, column=col, value=hdr)
        sc(c, bold=True, color='FFFFFF', fill=SUB, align='center')
    ws.row_dimensions[3].height = 18

    for i, skill in enumerate(SKILL_ORDER):
        row = 4 + i
        recs = recipes.get(skill, [])
        if not recs:
            continue

        req     = NEEDED_FOR_350.get(skill, 0)
        missing = len(recs)
        # Estimate learned = total - missing (approximate)
        total_est = req + missing  # rough
        learned_est = total_est - missing
        pct = round(learned_est / req * 100, 1) if req else 0
        still = max(0, req - learned_est)

        ready = buy = farm = 0
        for r in recs:
            status, _ = recipe_overall_status(r, inventory)
            if status == 'Ready Now!': ready += 1
            elif status == 'Buy/Craft Only': buy += 1
            else: farm += 1

        fill = ALT if i % 2 == 0 else WHT
        pf   = GRN if pct >= 100 else YLW if pct >= 70 else RED

        c1 = ws.cell(row=row, column=1, value=skill)
        c1.hyperlink = f'#{skill.replace(" ","").replace("/","")}!A1'
        c1.font = Font(name='Arial', size=10, color='1F4E79', underline='single')
        c1.fill = fill; c1.border = BDR
        c1.alignment = Alignment(horizontal='left', vertical='center')

        for col, val in [(2, missing), (3, req), (5, ready), (6, buy),
                         (7, farm), (8, still)]:
            c = ws.cell(row=row, column=col, value=val)
            sc(c, fill=fill, align='center')

        c4 = ws.cell(row=row, column=4, value=f'{pct}%')
        sc(c4, fill=pf, align='center', bold=True)

    # ── LEGEND ────────────────────────────────────────────────────────────────
    leg_row = 4 + len(SKILL_ORDER) + 1
    ws.merge_cells(f'A{leg_row}:H{leg_row}')
    c = ws[f'A{leg_row}']
    c.value = 'INGREDIENT LEGEND'
    sc(c, bold=True, color='FFFFFF', fill=SUB, align='center')

    legends = [
        (GRN,  '✓ Have it', 'Item is in your inventory'),
        (YLW,  'Buy from vendor', 'Available from a vendor/merchant'),
        (YLW2, 'Subcombine', 'Must be crafted first'),
        (RED,  'Farm/Drop', 'Must be hunted or looted'),
        (RED2, 'Forage/Fish', 'Must be foraged or fished'),
        (ALT,  'Check EQTraders', 'Source unknown — look up manually'),
    ]
    for j, (fill, label, desc) in enumerate(legends):
        r = leg_row + 1 + j
        c1 = ws.cell(row=r, column=1, value=label)
        sc(c1, fill=fill, bold=True)
        c2 = ws.cell(row=r, column=2, value=desc)
        ws.merge_cells(f'B{r}:H{r}')
        sc(c2, fill=WHT)

    # ── SKILL SHEETS ──────────────────────────────────────────────────────────
    for skill in SKILL_ORDER:
        recs = recipes.get(skill, [])
        if not recs:
            continue

        sname = skill.replace(' ', '').replace('/', '')[:31]
        ws2 = wb.create_sheet(sname)

        # Column widths
        ws2.column_dimensions['A'].width = 38  # recipe name
        ws2.column_dimensions['B'].width = 10  # trivial
        ws2.column_dimensions['C'].width = 14  # status
        ws2.column_dimensions['D'].width = 20  # expansion
        ws2.column_dimensions['E'].width = 20  # container
        # Ingredient columns (pairs: name + source)
        for col_idx in range(6, 22):
            ws2.column_dimensions[get_column_letter(col_idx)].width = 22

        # Title
        total_ingr_cols = 16  # 8 ingredients × 2 cols each
        total_cols = 5 + total_ingr_cols
        last_col = get_column_letter(total_cols)
        ws2.merge_cells(f'A1:{last_col}1')
        c = ws2['A1']
        c.value = f'{skill} — Missing Recipes (sorted: Ready → Buy → Farm)'
        sc(c, bold=True, color='FFFFFF', size=12, fill=HDR, align='center')
        ws2.row_dimensions[1].height = 22

        # Stats row
        req = NEEDED_FOR_350.get(skill, 0)
        ready_ct = sum(1 for r in recs if recipe_overall_status(r, inventory)[0] == 'Ready Now!')
        buy_ct   = sum(1 for r in recs if recipe_overall_status(r, inventory)[0] == 'Buy/Craft Only')
        farm_ct  = len(recs) - ready_ct - buy_ct

        ws2.merge_cells(f'A2:{last_col}2')
        c = ws2['A2']
        c.value = (f'{len(recs)} missing | {req} needed for 350 | '
                   f'{ready_ct} ready now | {buy_ct} buy only | {farm_ct} need farming')
        sc(c, italic=True, size=9, color='595959', fill=YLW)
        ws2.row_dimensions[2].height = 18

        # Headers
        base_hdrs = ['Recipe Name', 'Trivial', 'Status', 'Expansion', 'Container']
        ingr_hdrs = []
        for i in range(1, 9):
            ingr_hdrs += [f'Ingredient {i}', f'Source {i}']
        all_hdrs = base_hdrs + ingr_hdrs

        for col, hdr in enumerate(all_hdrs, 1):
            c = ws2.cell(row=3, column=col, value=hdr)
            sc(c, bold=True, color='FFFFFF', fill=SUB, align='center')
        ws2.row_dimensions[3].height = 18

        # Sort recipes: Ready → Buy → Farm, then by trivial
        def sort_key(r):
            status, _ = recipe_overall_status(r, inventory)
            try: triv = float(re.sub(r'[^\d.]', '', r.get('trivial', '999')))
            except: triv = 999
            order = {'Ready Now!': 0, 'Buy/Craft Only': 1, 'Need to Farm': 2}
            return (order.get(status, 3), triv)

        sorted_recs = sorted(recs, key=sort_key)

        for row_idx, r in enumerate(sorted_recs):
            row = 4 + row_idx
            even = row_idx % 2 == 0

            status, status_fill = recipe_overall_status(r, inventory)
            base_fill = (GRN if status == 'Ready Now!' else
                         YLW if status == 'Buy/Craft Only' else
                         RED if even else RED2)

            # Recipe name
            c1 = ws2.cell(row=row, column=1, value=r.get('name', ''))
            sc(c1, fill=base_fill)

            # Trivial
            c2 = ws2.cell(row=row, column=2, value=r.get('trivial', ''))
            sc(c2, fill=base_fill, align='center')

            # Status
            c3 = ws2.cell(row=row, column=3, value=status)
            status_color = ('375623' if status == 'Ready Now!' else
                            '7D6608' if status == 'Buy/Craft Only' else 'C00000')
            sc(c3, fill=status_fill, align='center', bold=True, color=status_color)

            # Expansion
            sc(ws2.cell(row=row, column=4, value=r.get('expansion', '')), fill=base_fill)

            # Container
            sc(ws2.cell(row=row, column=5, value=r.get('container', '')), fill=base_fill)

            # Ingredients
            for i in range(1, 9):
                ingr = r.get(f'ingredient_{i}', '').strip()
                src  = r.get(f'src_{i}', '').strip()
                col_ingr = 5 + (i - 1) * 2 + 1
                col_src  = col_ingr + 1

                have, status_text, ingr_fill = ingredient_status(ingr, src, inventory)

                c_ingr = ws2.cell(row=row, column=col_ingr, value=ingr)
                sc(c_ingr, fill=ingr_fill if ingr else base_fill)

                c_src = ws2.cell(row=row, column=col_src,
                                 value=status_text if ingr else '')
                src_color = ('375623' if have else
                             '7D6608' if 'vendor' in src.lower() or 'subcombine' in src.lower()
                             else 'C00000')
                sc(c_src, fill=ingr_fill if ingr else base_fill,
                   color=src_color if ingr else '000000',
                   bold=bool(ingr), size=9)

            ws2.row_dimensions[row].height = 30

    wb.save(OUTPUT_FILE)
    print(f"\n  Spreadsheet saved: {OUTPUT_FILE}")

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  EQ Tradeskill 350 Spreadsheet Builder")
    print("=" * 60)

    # Find inventory file
    inv_path = find_inventory_file()
    if inv_path:
        print(f"\nInventory file: {os.path.basename(inv_path)}")
    else:
        print("\nNo inventory file found — ingredient cross-reference disabled")
        print("Copy your CharName_Server-Inventory.txt to this folder for full features")

    # Load data
    print("\nLoading data...")
    inventory = load_inventory(inv_path)
    recipes   = load_recipes()

    # Build spreadsheet
    print("\nBuilding spreadsheet...")
    build_spreadsheet(recipes, inventory)

    # Summary
    total = sum(len(v) for v in recipes.values())
    print(f"\nDone!")
    print(f"  {total} missing recipes across {len(recipes)} tradeskills")
    print(f"  Output: {OUTPUT_FILE}")

if __name__ == '__main__':
    main()
